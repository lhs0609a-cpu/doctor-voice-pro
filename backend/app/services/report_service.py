"""
마케팅 리포트 생성 서비스
"""

import logging
from typing import Optional, Dict, List, Any
from datetime import datetime, date, timedelta
from calendar import monthrange
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, func
from sqlalchemy.orm import selectinload
import json
import io

from app.models.report import (
    MarketingReport, ReportSubscription,
    ReportType, ReportFormat, ReportStatus
)
from app.models.post import Post, PostStatus, PostAnalytics
from app.models.subscription import UsageLog

logger = logging.getLogger(__name__)


class ReportService:
    """마케팅 리포트 생성 서비스"""

    async def generate_report(
        self,
        db: AsyncSession,
        user_id: str,
        report_type: ReportType,
        period_start: date,
        period_end: date,
        title: Optional[str] = None,
    ) -> MarketingReport:
        """
        마케팅 리포트 생성

        Args:
            db: 데이터베이스 세션
            user_id: 사용자 ID
            report_type: 리포트 타입
            period_start: 시작일
            period_end: 종료일
            title: 리포트 제목

        Returns:
            생성된 MarketingReport
        """
        # 기본 제목 생성
        if not title:
            if report_type == ReportType.MONTHLY:
                title = f"{period_start.year}년 {period_start.month}월 마케팅 리포트"
            elif report_type == ReportType.WEEKLY:
                title = f"{period_start.strftime('%Y-%m-%d')} ~ {period_end.strftime('%Y-%m-%d')} 주간 리포트"
            else:
                title = f"{period_start.strftime('%Y-%m-%d')} ~ {period_end.strftime('%Y-%m-%d')} 리포트"

        # 리포트 생성
        report = MarketingReport(
            user_id=user_id,
            report_type=report_type,
            title=title,
            period_start=period_start,
            period_end=period_end,
            status=ReportStatus.GENERATING,
        )
        db.add(report)
        await db.commit()

        try:
            # 리포트 데이터 수집
            report_data = await self._collect_report_data(
                db, user_id, period_start, period_end
            )
            report.report_data = report_data
            report.status = ReportStatus.COMPLETED
            report.generated_at = datetime.utcnow()

        except Exception as e:
            logger.error(f"Failed to generate report: {e}")
            report.status = ReportStatus.FAILED
            report.error_message = str(e)

        await db.commit()
        await db.refresh(report)
        return report

    async def _collect_report_data(
        self,
        db: AsyncSession,
        user_id: str,
        period_start: date,
        period_end: date,
    ) -> Dict[str, Any]:
        """
        리포트 데이터 수집

        Args:
            db: 데이터베이스 세션
            user_id: 사용자 ID
            period_start: 시작일
            period_end: 종료일

        Returns:
            리포트 데이터 딕셔너리
        """
        start_dt = datetime.combine(period_start, datetime.min.time())
        end_dt = datetime.combine(period_end, datetime.max.time())

        # 1. 포스트 통계 수집
        posts_query = select(Post).where(
            and_(
                Post.user_id == user_id,
                Post.created_at >= start_dt,
                Post.created_at <= end_dt,
            )
        ).options(selectinload(Post.analytics))

        posts_result = await db.execute(posts_query)
        posts = posts_result.scalars().all()

        # 2. 요약 통계 계산
        total_posts = len(posts)
        published_posts = len([p for p in posts if p.status == PostStatus.PUBLISHED])
        draft_posts = len([p for p in posts if p.status == PostStatus.DRAFT])

        persuasion_scores = [p.persuasion_score for p in posts if p.persuasion_score]
        avg_persuasion_score = sum(persuasion_scores) / len(persuasion_scores) if persuasion_scores else 0

        total_views = sum(p.analytics.views if p.analytics else 0 for p in posts)
        total_inquiries = sum(p.analytics.inquiries if p.analytics else 0 for p in posts)

        summary = {
            "total_posts": total_posts,
            "published_posts": published_posts,
            "draft_posts": draft_posts,
            "publish_rate": round(published_posts / total_posts * 100, 1) if total_posts > 0 else 0,
            "avg_persuasion_score": round(avg_persuasion_score, 1),
            "total_views": total_views,
            "total_inquiries": total_inquiries,
            "avg_views_per_post": round(total_views / published_posts, 1) if published_posts > 0 else 0,
        }

        # 3. 설득력 점수 트렌드
        persuasion_trend = []
        current_date = period_start
        while current_date <= period_end:
            day_posts = [
                p for p in posts
                if p.created_at.date() == current_date and p.persuasion_score
            ]
            if day_posts:
                day_avg = sum(p.persuasion_score for p in day_posts) / len(day_posts)
                persuasion_trend.append({
                    "date": current_date.isoformat(),
                    "score": round(day_avg, 1),
                    "count": len(day_posts),
                })
            current_date += timedelta(days=1)

        # 4. 키워드 분석
        all_keywords = []
        for post in posts:
            if post.seo_keywords:
                if isinstance(post.seo_keywords, list):
                    all_keywords.extend(post.seo_keywords)
                elif isinstance(post.seo_keywords, dict) and 'keywords' in post.seo_keywords:
                    all_keywords.extend(post.seo_keywords['keywords'])

        keyword_counts = {}
        for kw in all_keywords:
            keyword_counts[kw] = keyword_counts.get(kw, 0) + 1

        top_keywords = sorted(
            keyword_counts.items(),
            key=lambda x: x[1],
            reverse=True
        )[:10]

        keyword_analysis = {
            "total_unique_keywords": len(keyword_counts),
            "top_keywords": [{"keyword": kw, "count": count} for kw, count in top_keywords],
        }

        # 5. 상위 포스트
        sorted_posts = sorted(
            [p for p in posts if p.persuasion_score],
            key=lambda x: x.persuasion_score,
            reverse=True
        )[:5]

        top_posts = [
            {
                "id": str(p.id),
                "title": p.title,
                "persuasion_score": p.persuasion_score,
                "status": p.status.value,
                "views": p.analytics.views if p.analytics else 0,
                "created_at": p.created_at.isoformat(),
            }
            for p in sorted_posts
        ]

        # 6. 개선 제안 생성
        recommendations = self._generate_recommendations(summary, keyword_analysis, top_posts)

        return {
            "summary": summary,
            "persuasion_trend": persuasion_trend,
            "keyword_analysis": keyword_analysis,
            "top_posts": top_posts,
            "recommendations": recommendations,
            "generated_at": datetime.utcnow().isoformat(),
        }

    def _generate_recommendations(
        self,
        summary: Dict,
        keyword_analysis: Dict,
        top_posts: List[Dict],
    ) -> List[Dict[str, str]]:
        """
        AI 기반 개선 제안 생성
        """
        recommendations = []

        # 발행률 기반 제안
        if summary["publish_rate"] < 70:
            recommendations.append({
                "type": "publishing",
                "priority": "high",
                "title": "발행률 개선 필요",
                "description": f"현재 발행률이 {summary['publish_rate']}%입니다. 작성한 글의 발행률을 높여 마케팅 효과를 극대화하세요.",
                "action": "임시저장된 글을 검토하고 발행을 진행하세요.",
            })

        # 설득력 점수 기반 제안
        if summary["avg_persuasion_score"] < 70:
            recommendations.append({
                "type": "quality",
                "priority": "medium",
                "title": "콘텐츠 품질 향상 권장",
                "description": f"평균 설득력 점수가 {summary['avg_persuasion_score']}점입니다. 상위 포스트의 작성 패턴을 참고하세요.",
                "action": "설득력 점수가 높은 글의 구조와 표현을 분석해보세요.",
            })

        # 발행 빈도 기반 제안
        if summary["total_posts"] < 8:
            recommendations.append({
                "type": "frequency",
                "priority": "medium",
                "title": "발행 빈도 증가 권장",
                "description": "월 8회 이상 발행 시 검색 노출이 크게 증가합니다.",
                "action": "주 2회 이상 정기적인 발행 스케줄을 설정하세요.",
            })

        # 키워드 다양성 제안
        if keyword_analysis["total_unique_keywords"] < 20:
            recommendations.append({
                "type": "keywords",
                "priority": "low",
                "title": "키워드 다양화 필요",
                "description": "다양한 키워드를 활용하여 더 많은 검색 유입을 확보하세요.",
                "action": "상위노출 분석 기능을 활용해 새로운 키워드를 발굴하세요.",
            })

        # 기본 긍정적 피드백
        if not recommendations:
            recommendations.append({
                "type": "positive",
                "priority": "info",
                "title": "우수한 성과를 유지하고 있습니다",
                "description": "현재 마케팅 활동이 잘 진행되고 있습니다. 이 추세를 유지하세요!",
                "action": "현재 전략을 유지하면서 새로운 키워드 발굴에 집중하세요.",
            })

        return recommendations

    async def get_reports(
        self,
        db: AsyncSession,
        user_id: str,
        report_type: Optional[ReportType] = None,
        limit: int = 20,
        offset: int = 0,
    ) -> List[MarketingReport]:
        """
        리포트 목록 조회
        """
        query = select(MarketingReport).where(
            MarketingReport.user_id == user_id
        ).order_by(MarketingReport.created_at.desc())

        if report_type:
            query = query.where(MarketingReport.report_type == report_type)

        query = query.offset(offset).limit(limit)
        result = await db.execute(query)
        return result.scalars().all()

    async def get_report(
        self,
        db: AsyncSession,
        report_id: str,
        user_id: Optional[str] = None,
    ) -> Optional[MarketingReport]:
        """
        리포트 상세 조회
        """
        query = select(MarketingReport).where(
            MarketingReport.id == report_id
        )

        if user_id:
            query = query.where(MarketingReport.user_id == user_id)

        result = await db.execute(query)
        return result.scalar_one_or_none()

    async def delete_report(
        self,
        db: AsyncSession,
        report_id: str,
        user_id: str,
    ) -> bool:
        """
        리포트 삭제
        """
        report = await self.get_report(db, report_id, user_id)
        if not report:
            return False

        await db.delete(report)
        await db.commit()
        return True

    async def generate_monthly_report(
        self,
        db: AsyncSession,
        user_id: str,
        year: int,
        month: int,
    ) -> MarketingReport:
        """
        월간 리포트 생성

        Args:
            db: 데이터베이스 세션
            user_id: 사용자 ID
            year: 연도
            month: 월

        Returns:
            생성된 MarketingReport
        """
        _, last_day = monthrange(year, month)
        period_start = date(year, month, 1)
        period_end = date(year, month, last_day)

        return await self.generate_report(
            db=db,
            user_id=user_id,
            report_type=ReportType.MONTHLY,
            period_start=period_start,
            period_end=period_end,
        )

    async def generate_weekly_report(
        self,
        db: AsyncSession,
        user_id: str,
        week_start: Optional[date] = None,
    ) -> MarketingReport:
        """
        주간 리포트 생성

        Args:
            db: 데이터베이스 세션
            user_id: 사용자 ID
            week_start: 주 시작일 (None이면 지난 주)

        Returns:
            생성된 MarketingReport
        """
        if not week_start:
            today = date.today()
            # 지난 주 월요일
            week_start = today - timedelta(days=today.weekday() + 7)

        period_end = week_start + timedelta(days=6)

        return await self.generate_report(
            db=db,
            user_id=user_id,
            report_type=ReportType.WEEKLY,
            period_start=week_start,
            period_end=period_end,
        )

    async def export_to_excel(
        self,
        db: AsyncSession,
        report_id: str,
        user_id: str,
    ) -> Optional[bytes]:
        """
        Excel 파일로 내보내기
        """
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl not installed")
            return None

        report = await self.get_report(db, report_id, user_id)
        if not report or not report.report_data:
            return None

        data = report.report_data
        wb = openpyxl.Workbook()

        # Summary 시트
        ws_summary = wb.active
        ws_summary.title = "요약"
        summary = data.get("summary", {})
        summary_rows = [
            ("항목", "값"),
            ("총 포스트", summary.get("total_posts", 0)),
            ("발행된 포스트", summary.get("published_posts", 0)),
            ("발행률", f"{summary.get('publish_rate', 0)}%"),
            ("평균 설득력 점수", summary.get("avg_persuasion_score", 0)),
            ("총 조회수", summary.get("total_views", 0)),
            ("총 문의수", summary.get("total_inquiries", 0)),
        ]
        for row in summary_rows:
            ws_summary.append(row)

        # Trend 시트
        ws_trend = wb.create_sheet("설득력 트렌드")
        ws_trend.append(("날짜", "평균 점수", "포스트 수"))
        for item in data.get("persuasion_trend", []):
            ws_trend.append((item["date"], item["score"], item["count"]))

        # Keywords 시트
        ws_keywords = wb.create_sheet("키워드 분석")
        ws_keywords.append(("키워드", "사용 횟수"))
        for kw in data.get("keyword_analysis", {}).get("top_keywords", []):
            ws_keywords.append((kw["keyword"], kw["count"]))

        # Top Posts 시트
        ws_posts = wb.create_sheet("상위 포스트")
        ws_posts.append(("제목", "설득력 점수", "상태", "조회수", "작성일"))
        for post in data.get("top_posts", []):
            ws_posts.append((
                post["title"],
                post["persuasion_score"],
                post["status"],
                post["views"],
                post["created_at"],
            ))

        # 바이트로 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    async def get_subscription(
        self,
        db: AsyncSession,
        user_id: str,
    ) -> Optional[ReportSubscription]:
        """
        리포트 자동 생성 구독 설정 조회
        """
        query = select(ReportSubscription).where(
            ReportSubscription.user_id == user_id
        )
        result = await db.execute(query)
        return result.scalar_one_or_none()

    async def update_subscription(
        self,
        db: AsyncSession,
        user_id: str,
        auto_monthly: Optional[bool] = None,
        auto_weekly: Optional[bool] = None,
        email_enabled: Optional[bool] = None,
        email_recipients: Optional[List[str]] = None,
        preferred_format: Optional[ReportFormat] = None,
    ) -> ReportSubscription:
        """
        리포트 자동 생성 구독 설정 업데이트
        """
        subscription = await self.get_subscription(db, user_id)

        if not subscription:
            subscription = ReportSubscription(user_id=user_id)
            db.add(subscription)

        if auto_monthly is not None:
            subscription.auto_monthly = auto_monthly
        if auto_weekly is not None:
            subscription.auto_weekly = auto_weekly
        if email_enabled is not None:
            subscription.email_enabled = email_enabled
        if email_recipients is not None:
            subscription.email_recipients = email_recipients
        if preferred_format is not None:
            subscription.preferred_format = preferred_format

        await db.commit()
        await db.refresh(subscription)
        return subscription


# Singleton instance
report_service = ReportService()
